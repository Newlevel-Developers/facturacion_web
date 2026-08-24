import openpyxl
from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from facturacion.models import Producto

@login_required
def exportar_productos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = ['ID / Código', 'Nombre', 'Precio Venta', 'Stock']
    ws.append(headers)

    # Filas con los datos reales usando getattr para evitar AttributeError
    productos = Producto.objects.all()
    for p in productos:
        # Intenta obtener 'codigo' o 'sku', si no existen usa el ID del producto
        codigo_identificador = getattr(p, 'codigo', getattr(p, 'sku', p.id))
        
        ws.append([
            codigo_identificador,
            p.nombre,
            float(p.precio_venta),
            p.stock
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Inventario_Productos.xlsx"'
    wb.save(response)
    return response


@login_required
def importar_productos_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']

        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls)')
            return redirect('productos')

        try:
            wb = openpyxl.load_workbook(archivo)
            sheet = wb.active

            creados = 0
            actualizados = 0

            with transaction.atomic():
                # Espera columnas: [Código/ID, Nombre, Precio Venta, Stock]
                for fila in sheet.iter_rows(min_row=2, values_only=True):
                    if not fila or not fila[1]: # Si la columna de Nombre está vacía, saltar
                        continue

                    identificador = str(fila[0]).strip() if fila[0] else None
                    nombre = str(fila[1]).strip()
                    precio_venta = Decimal(str(fila[2])) if len(fila) > 2 and fila[2] is not None else Decimal('0.00')
                    stock = int(fila[3]) if len(fila) > 3 and fila[3] is not None else 0

                    # Buscar o crear por nombre
                    producto, created = Producto.objects.update_or_create(
                        nombre=nombre,
                        defaults={
                            'precio_venta': precio_venta,
                            'stock': stock
                        }
                    )

                    if created:
                        creados += 1
                    else:
                        actualizados += 1

            messages.success(request, f'¡Éxito! Creados: {creados}, Actualizados: {actualizados}.')

        except Exception as e:
            messages.error(request, f'Error al importar: {str(e)}')

    return redirect('productos')